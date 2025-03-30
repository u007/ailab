import os
from openpyxl import load_workbook

file_path = "input.xlsx"
output_folder = "Extracted_Images"

wb = load_workbook(file_path, data_only=True)

if not os.path.exists(output_folder):
    os.mkdir(output_folder)
try:
    for sheet in wb.sheetnames:
        sheet_folder = os.path.join(output_folder, sheet)
        os.makedirs(sheet_folder, exist_ok=True)
        
        # Check if sheet exists and has images
        ws = wb[sheet]
        if not hasattr(ws, '_images'):
            print(f"No images found in sheet: {sheet}")
            continue
            
        for img in ws._images:  # OpenPyXL stores images in _images attribute
            try:
                img_path = os.path.join(sheet_folder, f"{img.ref}.png")
                with open(img_path, "wb") as f:
                    f.write(img._data)
                print(f"Successfully extracted image {img.ref} from sheet {sheet}")
            except Exception as e:
                print(f"Error extracting image {img.ref} from sheet {sheet}: {str(e)}")
                
except Exception as e:
    print(f"An error occurred during image extraction: {str(e)}")
print("Images extracted successfully.")