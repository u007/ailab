import os
import shutil
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Try to import xlrd for older Excel formats
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

def extract_images_from_xls(excel_path: str, output_folder: str):
    """Extract images from older Excel (.xls) files using xlrd"""
    if not XLRD_AVAILABLE:
        print("Warning: xlrd library not available. Cannot extract images from .xls files.")
        return []
    
    try:
        workbook = xlrd.open_workbook(excel_path, formatting_info=True)
        image_data = []
        
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            print(f"Processing sheet: {sheet_name} (XLS format)")
            
            # Get the sheet's drawings
            if hasattr(workbook, 'shape_types'):
                for drawing_idx, (sheet_rid, drawing) in enumerate(workbook.shape_types.items()):
                    if sheet_rid == sheet_idx and drawing.type == 3:  # Type 3 is for pictures
                        try:
                            img_name = f"{sheet_name}_{drawing_idx}"
                            img_path = os.path.join(output_folder, f"{img_name}.png")
                            # Check if image name already exists to avoid overwrites
                            counter = 1
                            while os.path.exists(img_path):
                                img_path = os.path.join(output_folder, f"{img_name}_{counter}.png")
                                counter += 1
                                
                            with open(img_path, 'wb') as f:
                                f.write(drawing.data)
                            image_data.append((img_name, img_path, None))
                        except Exception as e:
                            print(f"Error processing XLS image {drawing_idx} in sheet {sheet_name}: {str(e)}")
                            continue
        
        print(f"Extracted {len(image_data)} images from XLS file to {output_folder}")
        return image_data
    except Exception as e:
        print(f"Error extracting images from XLS file: {str(e)}")
        return []

def extract_images_from_excel(excel_path: str, output_folder: str):
    # Check if input file exists
    if not os.path.exists(excel_path):
        print(f"Error: Input Excel file '{excel_path}' does not exist")
        return

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"Error: Failed to load Excel file: {str(e)}")
        return
    
    print(f"Successfully loaded Excel file: {excel_path}")
    print(f"Found {len(wb.worksheets)} worksheet(s)")
    
    image_data = []
    try:
        for sheet in wb.worksheets:
            print(f"Processing sheet: {sheet.title}")
            # Access images through multiple methods to handle different Excel versions
            images = []
            
            # Method 1: Try accessing through _drawings if available
            if hasattr(sheet, '_drawings'):
                for drawing in sheet._drawings:
                    if hasattr(drawing, '_image_data') and drawing._image_data:
                        images.append(drawing)
                    elif hasattr(drawing, 'embedded') and drawing.embedded:
                        images.append(drawing)
            
            # Method 2: Try accessing through drawings property
            if hasattr(sheet, 'drawings'):
                for drawing in sheet.drawings:
                    if hasattr(drawing, 'image') and drawing.image:
                        images.append(drawing)
                    # Some Excel versions store image differently
                    elif hasattr(drawing, 'blip') and drawing.blip:
                        images.append(drawing)
            
            # Method 3: Check direct image collection
            if hasattr(sheet, '_images'):
                for img in sheet._images:
                    if hasattr(img, '_data') and img._data:
                        images.append(img)
                        
            # Method 4: Check for images in shapes collection
            if hasattr(sheet, '_shapes'):
                for shape in sheet._shapes:
                    if hasattr(shape, 'image') and shape.image:
                        images.append(shape)
                    elif hasattr(shape, '_image_data') and shape._image_data:
                        images.append(shape)
            
            # Method 5: Check for images in chart objects and hyperlinks in cells
            if hasattr(sheet, '_charts'):
                for chart in sheet._charts:
                    if hasattr(chart, 'picture') and chart.picture:
                        images.append(chart.picture)
            
            # Check for hyperlinks in cells that point to images
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink and hasattr(cell.hyperlink, 'target'):
                        target = cell.hyperlink.target
                        if target:
                            # Check if the hyperlink points to an image
                            image_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')
                            if any(target.lower().endswith(ext) for ext in image_extensions):
                                try:
                                    if target.startswith(('http://', 'https://')) or target.startswith('//'):
                                        if target.startswith('//'):
                                            target = 'https:' + target
                                        response = requests.get(target, timeout=10)
                                        if response.status_code == 200:
                                            img_data = response.content
                                            class ImageObject:
                                                def __init__(self, data):
                                                    self._image_data = data
                                            images.append(ImageObject(img_data))
                                            print(f"Successfully downloaded image from hyperlink: {target}")
                                except Exception as e:
                                    print(f"Error downloading image from hyperlink {target}: {str(e)}")
                                    continue
            
            # Method 6: Check for images in OpenXML format and linked images
            try:
                from openpyxl.packaging.relationship import RelationshipList
                import requests
                from urllib.parse import urljoin, urlparse
                
                if hasattr(sheet, '_rels') and sheet._rels:
                    for rel in sheet._rels.values():
                        # Check for image and hyperlink relationships
                        if rel.type and ('image' in rel.type.lower() or 'hyperlink' in rel.type.lower()):
                            try:
                                # Handle both local and remote paths
                                target = rel.target
                                # Check for common image extensions in the target
                                image_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')
                                is_image_url = any(target.lower().endswith(ext) for ext in image_extensions)
                                
                                # Handle URLs (both direct and relative)
                                if target.startswith(('http://', 'https://')) or target.startswith('//'):
                                    if target.startswith('//'):
                                        target = 'https:' + target
                                    # For external URLs
                                    try:
                                        if is_image_url:
                                            response = requests.get(target, timeout=10)
                                            if response.status_code == 200:
                                                img_data = response.content
                                                class ImageObject:
                                                    def __init__(self, data):
                                                        self._image_data = data
                                                images.append(ImageObject(img_data))
                                                print(f"Successfully downloaded linked image from {target}")
                                    except Exception as e:
                                        print(f"Error downloading linked image from {target}: {str(e)}")
                                else:
                                    # For local paths, check both in 'xl' folder and direct path
                                    possible_paths = [
                                        os.path.join(os.path.dirname(excel_path), 'xl', target),
                                        os.path.join(os.path.dirname(excel_path), target)
                                    ]
                                    
                                    for img_path in possible_paths:
                                        if os.path.exists(img_path) and any(img_path.lower().endswith(ext) for ext in image_extensions):
                                            try:
                                                with open(img_path, 'rb') as img_file:
                                                    img_data = img_file.read()
                                                    class ImageObject:
                                                        def __init__(self, data):
                                                            self._image_data = data
                                                    images.append(ImageObject(img_data))
                                                    print(f"Successfully loaded local image from {img_path}")
                                                    break
                                            except Exception as e:
                                                print(f"Error reading local image {img_path}: {str(e)}")
                                                continue
                            except Exception as e:
                                print(f"Error accessing image: {str(e)}")
            except ImportError:
                pass  # RelationshipList not available
            
            print(f"Found {len(images)} images in sheet {sheet.title}")
            for idx, image in enumerate(images):
                    try:
                        img_name = f"{sheet.title}_{idx}"  # Use sheet name and index as identifier
                        img_path = os.path.join(output_folder, f"{img_name}.png")
                        # Check if image name already exists to avoid overwrites
                        counter = 1
                        while os.path.exists(img_path):
                            img_path = os.path.join(output_folder, f"{img_name}_{counter}.png")
                            counter += 1
                        image_data.append((img_name, img_path, image))
                    except Exception as e:
                        print(f"Error processing image {idx} in sheet {sheet.title}: {str(e)}")
                        continue
    except Exception as e:
        print(f"Error accessing worksheet data: {str(e)}")
        return []
    
    # Sort images by name (row number)
    try:
        image_data.sort(key=lambda x: x[0])
    except Exception as e:
        print(f"Error sorting images: {str(e)}")
    
    try:
        for img_name, img_path, image in image_data:
            try:
                try:
                    if hasattr(image, '_image_data') and image._image_data:
                        with open(img_path, 'wb') as f:
                            f.write(image._image_data)
                    elif hasattr(image, 'embedded') and image.embedded:
                        with open(img_path, 'wb') as f:
                            f.write(image.embedded.image)
                    elif hasattr(image, '_data') and image._data:
                        with open(img_path, 'wb') as f:
                            f.write(image._data)
                    elif hasattr(image, 'blip') and hasattr(image.blip, 'embed') and image.blip.embed:
                        with open(img_path, 'wb') as f:
                            f.write(image.blip.embed)
                    elif hasattr(image, 'image') and image.image:
                        with open(img_path, 'wb') as f:
                            if hasattr(image.image, '_data') and image.image._data:
                                f.write(image.image._data)
                            elif hasattr(image.image, 'data') and image.image.data:
                                f.write(image.image.data)
                    elif hasattr(image, 'picture') and image.picture:
                        with open(img_path, 'wb') as f:
                            if hasattr(image.picture, '_data') and image.picture._data:
                                f.write(image.picture._data)
                            elif hasattr(image.picture, 'data') and image.picture.data:
                                f.write(image.picture.data)
                except Exception as e:
                    print(f"Warning: Could not save image {img_name} - {str(e)}")
                    continue
            except Exception as e:
                print(f"Error saving image {img_name}: {str(e)}")
                continue
    except Exception as e:
        print(f"Error processing images: {str(e)}")
        return
    
    print(f"Extracted and sorted {len(image_data)} images to {output_folder}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Error: Please provide the Excel file path")
        print("Usage: python extract.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_folder = "output_images"
    
    # Check file extension to determine which extraction method to use
    if excel_file.lower().endswith('.xls'):
        print("Detected .xls file format, using xlrd for extraction")
        if XLRD_AVAILABLE:
            extract_images_from_xls(excel_file, output_folder)
        else:
            print("Error: xlrd library is required for .xls files but not installed.")
            print("Install it using: pip install xlrd==1.2.0")
            sys.exit(1)
    else:
        print("Detected .xlsx file format, using openpyxl for extraction")
        extract_images_from_excel(excel_file, output_folder)
