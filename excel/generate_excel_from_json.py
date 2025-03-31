import json
import os
import requests
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

def download_image(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        return Image.open(BytesIO(response.content))
    except Exception as e:
        print(f"Error downloading image from {url}: {e}")
        return None

def resize_image(img, max_size=200):
    if img.width > max_size or img.height > max_size:
        ratio = min(max_size/img.width, max_size/img.height)
        new_width = int(img.width * ratio)
        new_height = int(img.height * ratio)
        return img.resize((new_width, new_height), Image.LANCZOS)
    return img
def generate_excel(filename):
    # Read JSON file
    json_path = os.path.join(os.path.dirname(__file__), filename)
    with open(json_path) as f:
        data = json.load(f)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = filename.replace('.json', '')
    
    # Set headers
    ws['A1'] = "Index"
    ws['B1'] = "Image"
    ws['C1'] = "URL"
    
    # Process each image URL
    for idx, url in enumerate(data['image_urls'], start=2):
        ws.cell(row=idx, column=1, value=idx-1)
        
        img = download_image(url)
        if img:
            img = resize_image(img)
            
            # Add image directly to Excel without temporary file
            try:
                # Save image to BytesIO buffer
                img_bytes = BytesIO()
                img.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                
                # Create Excel image from bytes
                excel_img = ExcelImage(img_bytes)
                ws.add_image(excel_img, f"B{idx}")
                
                # Add hyperlink to original URL
                ws.cell(row=idx, column=3).hyperlink = url
            except Exception as e:
                print(f"Error adding image to Excel: {e}")
            
    # Set column width
    ws.column_dimensions['B'].width = 26
    
    # Set row height for content rows
    for row in range(2, len(data['image_urls']) + 2):
        ws.row_dimensions[row].height = 60
    
    # Save Excel file
    json_filename = os.path.basename(json_path)
    excel_filename = json_filename.replace('.json', '.xlsx')
    output_path = os.path.join(os.path.dirname(__file__), excel_filename)
    wb.save(output_path)
    print(f"Excel file saved to {output_path}")

def main():
    generate_excel('sheet002.json')

if __name__ == "__main__":
    main()